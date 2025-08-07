import openpyxl
import mysql.connector
import os

# --- CẤU HÌNH DATABASE CỦA BẠN ---
# Thay thế bằng thông tin kết nối database của bạn
DB_CONFIG = {
    'host': 'timtruonghoc.mysql.pythonanywhere-services.com', # Thay bằng Host của bạn
    'user': 'timtruonghoc',            # Thay bằng Username database của bạn
    'password': 'admin@123',  # Thay bằng Password database của bạn
    'database': 'timtruonghoc$timtruonghocdb', # Thay bằng Database name của bạn
    'port': 3306 # Cổng mặc định của MySQL, hoặc 5432 cho PostgreSQL
}

# --- CẤU HÌNH FILE EXCEL CỦA BẠN ---
# Đảm bảo file Excel nằm trong cùng thư mục với script này
EXCEL_FILE_NAME = 'AllMajorOfAllSchool.xlsx' # Thay bằng tên file Excel của bạn
# Tên sheet chứa dữ liệu (nếu không phải sheet đầu tiên)
SHEET_NAME = 'Sheet1' # Thay bằng tên sheet của bạn

# --- CẤU HÌNH TÊN BẢNG VÀ CÁC CỘT TRONG DATABASE CỦA BẠN ---
# Chú ý: Thứ tự các cột phải khớp với thứ tự trong file Excel của bạn
TABLE_NAME = 'apptimtruonghoc_allmajorofallschool' # Thay 'app_name' bằng tên ứng dụng Django của bạn
DB_COLUMNS = [
    'all_major_id', 'name', 'training_duration', 'short_description',
    'tuition_fee_per_year', 'salary', 'field_id', 'opportunities','note',
    'job', 'suitable', 'program','tag'
]
# Các cột sẽ được cập nhật khi có trùng lặp 'all_major_id'
UPDATE_COLUMNS = [
    'name', 'training_duration', 'short_description', 'job', 'suitable',
    'program', 'salary', 'field_id', 'opportunities'
]


def import_all_major_data_from_excel():
    """
    Hàm đọc dữ liệu từ file Excel và chèn/cập nhật vào bảng allmajorofallschool.
    """
    excel_file_path = os.path.join(os.path.dirname(__file__), EXCEL_FILE_NAME)

    if not os.path.exists(excel_file_path):
        print(f"Lỗi: File Excel '{EXCEL_FILE_NAME}' không tìm thấy tại đường dẫn '{excel_file_path}'.")
        print("Vui lòng đảm bảo bạn đã đặt đúng tên file và file nằm trong cùng thư mục với script.")
        return

    try:
        # 1. Đọc dữ liệu từ file Excel
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active

        # Bỏ qua hàng tiêu đề
        header_skipped = False
        data_to_insert = []
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if not header_skipped:
                print(f"Hàng tiêu đề trong Excel: {row}")
                header_skipped = True
                continue

            processed_row = []
            # Lặp qua các cột đã định nghĩa để lấy dữ liệu
            for i, col_name in enumerate(DB_COLUMNS):
                excel_value = row[i]

                # Xử lý giá trị None/chuỗi rỗng
                if excel_value == "" or excel_value is None:
                    processed_row.append(None)
                else:
                    # Chuyển đổi về string và loại bỏ khoảng trắng
                    processed_row.append(str(excel_value).strip())

            # Chỉ thêm vào nếu hàng không rỗng
            if any(processed_row):
                data_to_insert.append(tuple(processed_row))

        if not data_to_insert:
            print("Không tìm thấy dữ liệu nào để chèn sau khi bỏ qua tiêu đề.")
            return

        print(f"Đã đọc {len(data_to_insert)} hàng dữ liệu từ Excel.")

        # 2. Kết nối đến Database và chèn dữ liệu
        conn = None
        cursor = None
        try:
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor()

            # Tạo câu lệnh SQL INSERT động
            columns_str = ', '.join(DB_COLUMNS)
            placeholders = ', '.join(['%s'] * len(DB_COLUMNS))

            # Tạo phần ON DUPLICATE KEY UPDATE
            update_str = ', '.join([f"{col}=VALUES({col})" for col in UPDATE_COLUMNS])

            # Sử dụng INSERT ... ON DUPLICATE KEY UPDATE để tránh lỗi trùng lặp all_major_id
            insert_query = f"""
                INSERT INTO {TABLE_NAME} ({columns_str}) VALUES ({placeholders})
                ON DUPLICATE KEY UPDATE {update_str}
            """

            print(f"Đang chèn/cập nhật {len(data_to_insert)} hàng dữ liệu vào bảng '{TABLE_NAME}'...")
            cursor.executemany(insert_query, data_to_insert)
            conn.commit()

            print(f"✔ Nhập liệu thành công {cursor.rowcount} hàng dữ liệu vào bảng '{TABLE_NAME}'.")

        except mysql.connector.Error as err:
            print(f"Lỗi Database: {err}")
            if conn:
                conn.rollback()
        except Exception as e:
            print(f"Lỗi không xác định: {e}")
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()
                print("Đã đóng kết nối Database.")

    except Exception as e:
        print(f"Lỗi khi đọc file Excel: {e}")
        print("Vui lòng kiểm tra lại tên sheet và định dạng file Excel của bạn.")

if __name__ == "__main__":
    import_all_major_data_from_excel()
