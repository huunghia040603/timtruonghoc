import openpyxl
import mysql.connector # Hoặc import psycopg2 nếu dùng PostgreSQL
import os

# --- CẤU HÌNH DATABASE CỦA BẠN ---
DB_CONFIG = {
    'host': 'webtimtruong.mysql.pythonanywhere-services.com',
    'user': 'webtimtruong',
    'password': '15django432',
    'database': 'webtimtruong$timtruongdb',
    'port': 3306
}

# --- CẤU HÌNH FILE EXCEL CỦA BẠN ---
# Đảm bảo file Excel nằm trong cùng thư mục với script này
EXCEL_FILE_NAME = 'FieldGroup1.xlsx' # Thay bằng tên file Excel của bạn
# Tên sheet chứa dữ liệu (nếu không phải sheet đầu tiên)
SHEET_NAME = 'Sheet1' # Thay bằng tên sheet của bạn

# --- CẤU HÌNH TÊN BẢNG VÀ CÁC CỘT TRONG DATABASE CỦA BẠN ---
# Đảm bảo thứ tự cột trong list này khớp với thứ tự dữ liệu bạn đọc từ Excel
TABLE_NAME = 'apptimtruonghoc_fieldgroup' # Thay bằng tên bảng của bạn (vd: apptimtruonghoc_fieldgroup)
DB_COLUMNS = ['field_id','name','description','cover',
]

def import_fieldgroup_data_from_excel():
    excel_file_path = os.path.join(os.path.dirname(__file__), EXCEL_FILE_NAME)

    if not os.path.exists(excel_file_path):
        print(f"Lỗi: File Excel '{EXCEL_FILE_NAME}' không tìm thấy tại đường dẫn '{excel_file_path}'.")
        print("Vui lòng đảm bảo bạn đã đặt đúng tên file và file nằm trong cùng thư mục với script.")
        return

    try:
        # 1. Đọc dữ liệu từ file Excel
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active

        # Bỏ qua hàng tiêu đề
        header_skipped = False
        data_to_insert = []
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if not header_skipped:
                print(f"Hàng tiêu đề trong Excel: {row}")
                header_skipped = True
                continue

            processed_row = []
            for i, col_name in enumerate(DB_COLUMNS):
                excel_value = row[i]

                # Xử lý giá trị None/chuỗi rỗng
                if excel_value == "" or excel_value is None:
                    processed_row.append(None)
                else:
                    processed_row.append(str(excel_value).strip())

            data_to_insert.append(tuple(processed_row))

        if not data_to_insert:
            print("Không tìm thấy dữ liệu nào để chèn sau khi bỏ qua tiêu đề.")
            return

        print(f"Đã đọc {len(data_to_insert)} hàng dữ liệu từ Excel.")

        # 2. Kết nối đến Database và chèn dữ liệu
        conn = None
        cursor = None
        try:
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor()

            # Tạo câu lệnh SQL INSERT động
            columns_str = ', '.join(DB_COLUMNS)
            placeholders = ', '.join(['%s'] * len(DB_COLUMNS))

            # Sử dụng INSERT ... ON DUPLICATE KEY UPDATE để tránh lỗi trùng lặp field_id
            insert_query = f"""
                INSERT INTO {TABLE_NAME} ({columns_str}) VALUES ({placeholders})
                ON DUPLICATE KEY UPDATE
                name=VALUES(name), 
                description=VALUES(description),
                cover=VALUES(cover)
            """

            print(f"Đang chèn/cập nhật {len(data_to_insert)} hàng dữ liệu vào bảng '{TABLE_NAME}'...")
            cursor.executemany(insert_query, data_to_insert)
            conn.commit()

            print(f"✔ Nhập liệu thành công {cursor.rowcount} hàng dữ liệu vào bảng '{TABLE_NAME}'.")

        except mysql.connector.Error as err:
            print(f"Lỗi Database: {err}")
            if conn:
                conn.rollback()
        except Exception as e:
            print(f"Lỗi không xác định: {e}")
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()
                print("Đã đóng kết nối Database.")

    except Exception as e:
        print(f"Lỗi khi đọc file Excel: {e}")
        print("Vui lòng kiểm tra lại tên sheet và định dạng file Excel của bạn.")

if __name__ == "__main__":
    import_fieldgroup_data_from_excel()
