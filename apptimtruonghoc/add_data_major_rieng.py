import openpyxl
import mysql.connector # Sử dụng mysql.connector cho MySQL
import os

# --- CẤU HÌNH DATABASE CỦA BẠN ---
DB_CONFIG = {
    'host': 'timtruonghoc.mysql.pythonanywhere-services.com', # Thay bằng Host của bạn
    'user': 'timtruonghoc',            # Thay bằng Username database của bạn
    'password': 'admin@123',  # Thay bằng Password database của bạn
    'database': 'timtruonghoc$timtruonghocdb', # Thay bằng Database name của bạn
    'port': 3306 # Cổng mặc định của MySQL, hoặc 5432 cho PostgreSQL
}

# --- CẤU HÌNH FILE EXCEL CỦA BẠN ---
# Đảm bảo file Excel nằm trong cùng thư mục với script này
EXCEL_FILE_NAME = 'MajorRiengTruong.xlsx' # Thay bằng tên file Excel của bạn
# Tên sheet chứa dữ liệu (nếu không phải sheet đầu tiên)
SHEET_NAME = 'Sheet1' # Thay bằng tên sheet của bạn

# --- CẤU HÌNH TÊN BẢNG VÀ CÁC CỘT TRONG DATABASE CỦA BẠN ---
# Đảm bảo thứ tự cột trong list này khớp với thứ tự dữ liệu bạn đọc từ Excel
TABLE_NAME = 'apptimtruonghoc_major' # Tên bảng Major trong Django

# Các cột trong bảng database Major, bao gồm cả school_id cho ForeignKey
# Đã loại bỏ 'tags' khỏi danh sách cột
DB_COLUMNS = [
    'major_id', 'name', 'description', 'entry_requirement',
    'min_tuition_fee_per_year', 'max_tuition_fee_per_year', 'status','tags', 'school_id'
]

# --- CẤU HÌNH ÁNH XẠ CỘT TỪ EXCEL SANG TRƯỜNG TRONG DATABASE ---
# Key là tên cột trong file Excel (tiêu đề), Value là tên trường trong database
# Đã loại bỏ 'tags' khỏi ánh xạ
EXCEL_TO_DB_COLUMN_MAP = {
    'Mã ngành': 'major_id',
    'Tên ngành': 'name',
    'Mô tả ngành (Điểm khác biệt so với đào tạo ở trường khác)': 'description',
    'Phương thức xét tuyển': 'entry_requirement',
    'Học phí tối thiểu': 'min_tuition_fee_per_year',
    'Học phí tối đa': 'max_tuition_fee_per_year',
    'Ngành còn hoạt động?': 'status',
    'tags': 'tags',
    'mã trường': 'school_id',
     # Ánh xạ cột 'mã trường' trong Excel sang 'school_id' trong DB
}


def import_major_data_from_excel():
    """
    Hàm này đọc dữ liệu từ file Excel và nhập/cập nhật trực tiếp vào database
    sử dụng mysql.connector.
    """
    excel_file_path = os.path.join(os.path.dirname(__file__), EXCEL_FILE_NAME)

    if not os.path.exists(excel_file_path):
        print(f"Lỗi: File Excel '{EXCEL_FILE_NAME}' không tìm thấy tại đường dẫn '{excel_file_path}'.")
        print("Vui lòng đảm bảo bạn đã đặt đúng tên file và file nằm trong cùng thư mục với script.")
        return

    try:
        # 1. Đọc dữ liệu từ file Excel
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active

        # Đọc hàng tiêu đề để tạo ánh xạ cột
        header = [cell.value for cell in sheet[1]]
        print(f"Hàng tiêu đề trong Excel: {header}")

        # Tạo ánh xạ từ tên cột Excel sang chỉ số cột
        excel_column_indices = {}
        for excel_col_name, db_field_name in EXCEL_TO_DB_COLUMN_MAP.items():
            try:
                # Tìm chỉ số của tên cột Excel trong hàng tiêu đề
                excel_column_indices[db_field_name] = header.index(excel_col_name)
            except ValueError:
                print(f"Lỗi: Không tìm thấy cột '{excel_col_name}' trong file Excel.")
                print("Vui lòng kiểm tra lại tên cột trong file Excel của bạn khớp với EXCEL_TO_DB_COLUMN_MAP.")
                return # Thoát nếu một cột bắt buộc bị thiếu

        data_to_insert = []
        # Bắt đầu đọc từ hàng thứ 2 (min_row=2) để bỏ qua tiêu đề
        for row_idx, row_data_tuple in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            row_number_in_excel = row_idx + 2 # Hàng trên Excel (có tính tiêu đề)

            processed_row_values = [None] * len(DB_COLUMNS) # Khởi tạo với None

            # Lấy dữ liệu từ hàng Excel và ánh xạ vào processed_row_values theo thứ tự DB_COLUMNS
            for db_field_name, col_index in excel_column_indices.items():
                if col_index < len(row_data_tuple): # Đảm bảo chỉ số hợp lệ
                    excel_value = row_data_tuple[col_index]

                    # Xử lý giá trị None/chuỗi rỗng
                    if excel_value == "" or excel_value is None:
                        processed_value = None
                    else:
                        processed_value = str(excel_value).strip()

                        # Xử lý học phí: loại bỏ dấu chấm phân cách hàng nghìn
                        if db_field_name in ['min_tuition_fee_per_year', 'max_tuition_fee_per_year']:
                            processed_value = processed_value.replace('.', '')

                    # Gán giá trị vào đúng vị trí trong processed_row_values
                    # Tìm vị trí của db_field_name trong DB_COLUMNS
                    try:
                        db_col_pos = DB_COLUMNS.index(db_field_name)
                        processed_row_values[db_col_pos] = processed_value
                    except ValueError:
                        # Trường không có trong DB_COLUMNS (ví dụ: 'tags' đã bị loại bỏ)
                        pass # Bỏ qua nếu trường không có trong danh sách cột DB_COLUMNS

            # Kiểm tra các trường bắt buộc (major_id, name, school_id)
            major_id = processed_row_values[DB_COLUMNS.index('major_id')]
            name = processed_row_values[DB_COLUMNS.index('name')]
            school_id = processed_row_values[DB_COLUMNS.index('school_id')]

            if not major_id or not name or not school_id:
                print(f"Bỏ qua hàng {row_number_in_excel} (trên Excel) do thiếu Mã ngành, Tên ngành hoặc Mã trường.")
                continue

            # Đã loại bỏ logic xử lý cột 'tags'

            data_to_insert.append(tuple(processed_row_values))


        if not data_to_insert:
            print("Không tìm thấy dữ liệu nào để chèn sau khi bỏ qua tiêu đề.")
            return

        print(f"Đã đọc {len(data_to_insert)} hàng dữ liệu từ Excel.")

        # 2. Kết nối đến Database và chèn dữ liệu
        conn = None
        cursor = None
        try:
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor()

            # Tạo câu lệnh SQL INSERT động
            columns_str = ', '.join(DB_COLUMNS)
            placeholders = ', '.join(['%s'] * len(DB_COLUMNS))

            # Tạo phần UPDATE cho ON DUPLICATE KEY UPDATE
            update_parts = []
            # Không cập nhật major_id và school_id vì chúng là một phần của khóa duy nhất
            for col in DB_COLUMNS:
                if col not in ['major_id', 'school_id']: # Không cập nhật các cột khóa chính/duy nhất
                    update_parts.append(f"{col}=VALUES({col})")
            update_str = ', '.join(update_parts)

            # Sử dụng INSERT ... ON DUPLICATE KEY UPDATE để tránh lỗi trùng lặp
            insert_query = f"""
                INSERT INTO {TABLE_NAME} ({columns_str}) VALUES ({placeholders})
                ON DUPLICATE KEY UPDATE
                {update_str}
            """

            print(f"Đang chèn/cập nhật {len(data_to_insert)} hàng dữ liệu vào bảng '{TABLE_NAME}'...")
            cursor.executemany(insert_query, data_to_insert)
            conn.commit()

            print(f"✔ Nhập liệu thành công {cursor.rowcount} hàng dữ liệu vào bảng '{TABLE_NAME}'.")

        except mysql.connector.Error as err:
            print(f"Lỗi Database: {err}")
            if conn:
                conn.rollback() # Hoàn tác các thay đổi nếu có lỗi
        except Exception as e:
            print(f"Lỗi không xác định: {e}")
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()
                print("Đã đóng kết nối Database.")

    except Exception as e:
        print(f"Lỗi khi đọc file Excel: {e}")
        print("Vui lòng kiểm tra lại tên sheet và định dạng file Excel của bạn.")

if __name__ == "__main__":
    import_major_data_from_excel()
