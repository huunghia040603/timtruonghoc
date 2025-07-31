import openpyxl
import mysql.connector # Hoặc import psycopg2 nếu dùng PostgreSQL
import os # Để lấy thông tin đường dẫn

# --- CẤU HÌNH DATABASE CỦA BẠN ---
DB_CONFIG = {
    'host': 'webtimtruong.mysql.pythonanywhere-services.com', # Thay bằng Host của bạn
    'user': 'webtimtruong',            # Thay bằng Username database của bạn
    'password': '15django432',  # Thay bằng Password database của bạn
    'database': 'webtimtruong$timtruongdb', # Thay bằng Database name của bạn
    'port': 3306 # Cổng mặc định của MySQL, hoặc 5432 cho PostgreSQL
}

# --- CẤU HÌNH FILE EXCEL CỦA BẠN ---
# Đảm bảo file Excel nằm trong cùng thư mục với script này
EXCEL_FILE_NAME = 'Book.xlsx' # Thay bằng tên file Excel của bạn
# Tên sheet chứa dữ liệu (nếu không phải sheet đầu tiên)
SHEET_NAME = 'Sheet1' # Thay bằng tên sheet của bạn

# --- CẤU HÌNH TÊN BẢNG VÀ CÁC CỘT TRONG DATABASE CỦA BẠN ---
# Đảm bảo thứ tự cột trong list này khớp với thứ tự dữ liệu bạn đọc từ Excel
TABLE_NAME = 'apptimtruonghoc_school' # Thay bằng tên bảng của bạn (vd: appname_modelname)
DB_COLUMNS = [
            'name_vn', 'name_en', 'short_code','admission_code', 'logo',
            'cover_photo', 'established_year', 'address','country',
            'school_type', 'website_url', 'phone_number','start','end',
            'school_level','email','map_link','quota_per_year',
            'introduction','socialmedialink','benchmark_min', 'benchmark_max','tag',
            'registration'

        ]

def import_data_from_excel():
    excel_file_path = os.path.join(os.path.dirname(__file__), EXCEL_FILE_NAME)

    if not os.path.exists(excel_file_path):
        print(f"Lỗi: File Excel '{EXCEL_FILE_NAME}' không tìm thấy tại đường dẫn '{excel_file_path}'.")
        print("Vui lòng đảm bảo bạn đã tải file Excel lên PythonAnywhere và đặt đúng tên.")
        return

    try:
        # 1. Đọc dữ liệu từ file Excel
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active # Lấy sheet theo tên hoặc sheet đầu tiên

        # Bỏ qua hàng tiêu đề nếu có (ví dụ: bắt đầu từ hàng thứ 2)
        header_skipped = False
        data_to_insert = []
        for row in sheet.iter_rows(min_row=1, values_only=True): # Bắt đầu từ hàng 1 để đọc header
            if not header_skipped:
                # Nếu hàng đầu tiên là tiêu đề, bạn có thể in ra để kiểm tra
                print(f"Hàng tiêu đề trong Excel: {row}")
                header_skipped = True
                continue # Bỏ qua hàng tiêu đề

            processed_row = []
            for i, col_name in enumerate(DB_COLUMNS):
                excel_value = row[i] # Lấy giá trị từ cột thứ i trong hàng Excel

                # Xử lý giá trị None/chuỗi rỗng cho các trường NULLABLE
                if excel_value == "" or excel_value is None:
                    # For unique fields like 'short_code', it's often better to insert NULL
                    # instead of an empty string if it's truly meant to be empty.
                    # MySQL treats multiple NULLs in a UNIQUE key as distinct, but multiple '' as duplicates.
                    if col_name == 'short_code': # Apply specific handling for short_code if it's the problem
                        processed_row.append(None) # Insert NULL for empty short_code
                    else:
                        processed_row.append(None)
                else:
                    # Chuyển đổi kiểu dữ liệu nếu cần
                    if col_name in ['established_year', 'quota_per_year']:
                        try:
                            processed_row.append(int(excel_value))
                        except (ValueError, TypeError):
                            processed_row.append(None) # Hoặc giá trị mặc định/báo lỗi
                    elif col_name in ['admission_score_threshold']:
                        try:
                            processed_row.append(float(excel_value))
                        except (ValueError, TypeError):
                            processed_row.append(None) # Hoặc giá trị mặc định/báo lỗi
                    else:
                        str_value = str(excel_value).strip()
                        # Explicitly handle empty string after stripping for 'short_code'
                        if col_name == 'short_code' and str_value == '':
                            processed_row.append(None) # Treat stripped empty string as NULL
                        else:
                            processed_row.append(str_value) # Đảm bảo là chuỗi và bỏ khoảng trắng thừa

            data_to_insert.append(tuple(processed_row))

        if not data_to_insert:
            print("Không tìm thấy dữ liệu nào để chèn sau khi bỏ qua tiêu đề.")
            return

        print(f"Đã đọc {len(data_to_insert)} hàng dữ liệu từ Excel.")
        # print("Dữ liệu mẫu từ Excel (hàng đầu tiên):", data_to_insert[0]) # In ra để kiểm tra

        # 2. Kết nối đến Database và chèn dữ liệu
        conn = None
        cursor = None
        try:
            conn = mysql.connector.connect(**DB_CONFIG) # Dùng mysql.connector
            # Hoặc: conn = psycopg2.connect(**DB_CONFIG) # Nếu dùng PostgreSQL
            cursor = conn.cursor()

            # Tạo câu lệnh SQL INSERT động dựa trên tên bảng và các cột
            columns_str = ', '.join(DB_COLUMNS)
            placeholders = ', '.join(['%s'] * len(DB_COLUMNS)) # Dùng %s cho MySQL/psycopg2

            insert_query = f"INSERT INTO {TABLE_NAME} ({columns_str}) VALUES ({placeholders})"

            print(f"Đang chèn {len(data_to_insert)} hàng dữ liệu vào bảng '{TABLE_NAME}'...")
            cursor.executemany(insert_query, data_to_insert)
            conn.commit() # Lưu các thay đổi vào database

            print(f"✔ Nhập liệu thành công {cursor.rowcount} hàng dữ liệu vào bảng '{TABLE_NAME}'.")

        except mysql.connector.Error as err: # Hoặc psycopg2.Error nếu dùng PostgreSQL
            print(f"Lỗi Database: {err}")
            if conn:
                conn.rollback() # Hoàn tác các thay đổi nếu có lỗi
        except Exception as e:
            print(f"Lỗi không xác định: {e}")
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()
                print("Đã đóng kết nối Database.")

    except Exception as e:
        print(f"Lỗi khi đọc file Excel: {e}")
        print("Vui lòng kiểm tra lại tên sheet và định dạng file Excel của bạn.")

if __name__ == "__main__":
    import_data_from_excel()



